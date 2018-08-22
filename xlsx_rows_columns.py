#! python3

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']  # before wb.get_sheet_by_name('Sheet1')

# sheet['A1':'C7'] is this:
# ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, ....

for row_of_cell_objects in sheet['A1':'C7']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print('--- END OF ROW ---')

print('-------------------------------------------------------------------------------------------')

for row_num in range(1, sheet.max_row + 1):
    print(f'--- ROW NUMBER {row_num} ---')
    for col_num in range(1, sheet.max_column + 1):
        print(sheet.cell(row=row_num, column=col_num).value)

# print(tuple(sheet['A1':'C7']))
"""
Cookbook:

- Import the openpyxl module.
- Call the openpyxl.load_workbook() function. OK
- Get a Workbook object.
- Read the active member variable 'wb.active' or 'wb[sheet]'.
- Get a Worksheet object.
- Use indexing or the cell() sheet method with row and column keyword arguments 
  (sheet.columns or sheet.rows after converted to a tuple).
- Get a Cell object.
- Read the Cell object’s value attribute (cell_obj.value)

To read a value of a cell there are three ways - sheet.cell(1, 1).value
                                               - sheet.cell(row=1, column=1).value
                                               - sheet[a1].value
"""


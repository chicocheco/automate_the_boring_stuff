#! python3
# xlsx_read_census.py - Tabulates population and number of census tracts for each county.

# To read a value of a cell there are three ways - sheet.cell(1, 1).value
#                                                - sheet.cell(row=1, column=1).value
#                                                - sheet[a1].value

import openpyxl
import pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']

# Design a structure of email_data - {'state': {'census': {'pop': 3141, 'tracts': 1}, {...}...}...}
# examples: countyData['AK']['Anchorage']['pop'] = 291826
#           countyData['AK']['Anchorage']['tracts'] = 55
county_data = {}
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has email_data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # Make sure the key for this state exists.
    county_data.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    county_data[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    county_data[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')
result_file = open('census2010.py', 'w')
result_file.write('all_data = ' + pprint.pformat(county_data))  # Formatted as valid Python code.
result_file.close()
print('Done.')

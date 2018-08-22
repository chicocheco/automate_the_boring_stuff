
from openpyxl.utils import get_column_letter, column_index_from_string  # before it was 'openpyxl.cell'
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))



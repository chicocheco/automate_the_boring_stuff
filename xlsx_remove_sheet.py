import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)

# Create two sheets more. The second argument is its index.
wb.create_sheet('First Sheet', 0)
wb.create_sheet('Last Sheet')
print(wb.sheetnames)

# Remove the middle sheet.
wb.remove(wb['Sheet'])  # old: wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
print(wb.sheetnames)

# The remove() method takes a Worksheet object, not a string of the sheet name, as its argument.

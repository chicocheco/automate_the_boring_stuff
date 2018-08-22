#! python3

import openpyxl

input_wb = openpyxl.load_workbook('example.xlsx')
input_sheet = input_wb.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

# Nested "for" loops to read in the spreadsheet’s email_data into a list of lists email_data structure.

for row_num in range(1, input_sheet.max_row + 1):
    for col_num in range(1, input_sheet.max_column + 1):
        output_sheet.cell(row=col_num, column=row_num).value = input_sheet.cell(row=row_num, column=col_num).value

output_wb.save('example_inverted.xlsx')
output_wb.close()
input_wb.close()



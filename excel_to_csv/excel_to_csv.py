#! python3
# excel_to_csv.py - Converts Excel spreadsheets to CSV files, naming them by work sheets.

import openpyxl
import os
import csv
import logging

#  logging.disable(logging.CRITICAL) to disable logging!!!
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program')

for excel_file in os.listdir('.'):
    logging.debug('File: ' + excel_file)
    if not excel_file.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        logging.debug('Name of the sheet to convert: ' + sheet_name)
        sheet_name_fix = sheet_name.replace(' ', '_').lower()
        csv_filename = os.path.splitext(excel_file)[0] + '_' + sheet_name_fix + '.csv'
        output_csv = open(csv_filename, 'w', newline='')
        output_csv_writer = csv.writer(output_csv)

        # Loop through every row in the sheet and create a list of sublists.
        sheet_data = []
        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            sheet_data.append(row_data)

            for col_num in range(1, sheet.max_column + 1):
                cell_content = str(sheet.cell(row=row_num, column=col_num).value)
                row_data.append(cell_content)

        logging.debug('Row email_data created: ' + str(sheet_data))

        # len(sheet_data) is 201. If I add + 1, it would be 202, what is wrong.
        for csv_row in range(0, len(sheet_data)):
            output_csv_writer.writerow(sheet_data[csv_row])
        logging.debug('Name of the new CSV: ' + csv_filename)
        logging.debug('====================================================')
        output_csv.close()

logging.debug('End of program')

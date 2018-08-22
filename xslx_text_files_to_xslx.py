#! python3
# xlsx_text_to_xlsx.py - Search for .txt files and copy their content
#                        in individual columns of a new excel spreahsheet.

import openpyxl
import os


def text_to_spreadsheet(folder):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Create a list of text files.
    txt_files = []
    for root, subfolds, files in os.walk(folder):
        for file in files:
            extensions = '.txt'
            if file.lower().endswith(extensions):
                txt_files.append(file)

    # Fill out the spreadsheet with the content of text files.
    file_num = 0
    for txt_file in txt_files:
        file_obj = open(txt_file)
        col = file_obj.read().splitlines()
        col = [value for value in col if value != '']
        col.insert(0, 'to skip')
        col.insert(1, txt_file)
        file_num += 1
        for row_num in range(1, len(col)):
            sheet.cell(row=row_num, column=file_num).value = col[row_num]
            # sheet.column_dimensions[openpyxl.utils.get_column_letter(file_num)].width = 50

    print(f'I have found and processed the following {len(txt_files)} text files:\n{txt_files}')
    wb.save('text_files_to_spreadsheet.xlsx')
    wb.close()


text_to_spreadsheet(input('Enter a name of a folder (subfolders included) to search for .txt files '
                          'and copy their content in individual columns of a new excel spreahsheet'
                          ' (text_files_to_spreadsheet.xlsx): \n'))

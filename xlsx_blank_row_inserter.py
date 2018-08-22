#! python3
# xlsx_blank_row_inserter.py - Insert N blanks rows starting from M'th row.
# Author: Priyank Jain
# WARNING: It is very slow with spreadsheets of thousands rows.
#####################################################################
import openpyxl
import os
import sys
print("Launched.")
if len(sys.argv) < 4:
    print('Usage: python3 blank_row_inserter.py m n excel_file_path.')
    sys.exit(-1)
m = None
n = None
excel_file_path = None

try:
    m = int(sys.argv[1])
    n = int(sys.argv[2])
    excel_file_path = sys.argv[3]
except ValueError as e:
    print('Both m and n should be integers.')
    sys.exit(-2)
if not os.path.exists(excel_file_path):
    print('The specified file does not exist.')
    sys.exit(-3)

input_wb = openpyxl.load_workbook(excel_file_path)
input_sheet = input_wb.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active

print('Working...')
for row_num in range(1, input_sheet.max_row + 1):
    for col_num in range(1, input_sheet.max_column + 1):
        if row_num < m:
            output_sheet.cell(row=row_num, column=col_num).value = input_sheet.cell(row=row_num, column=col_num).value
        else:
            output_sheet.cell(row=row_num + n, column=col_num).value = \
                input_sheet.cell(row=row_num, column=col_num).value

print('Done!')
file_name, file_extension = os.path.splitext(excel_file_path)
output_wb.save(f'{file_name}_with_blanks{file_extension}')
input_wb.save(excel_file_path)

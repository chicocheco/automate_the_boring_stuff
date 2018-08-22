#! python3
# multiplication_table.py - Creates an N x N multiplication table in an Excel spreadsheet and save it on the desktop.

import openpyxl
import sys
from pathlib import Path

number = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
bold_font = openpyxl.styles.Font(bold=True)

# The second argument of the 'range()' is +2 because the first cell has to be empty.
for row_num in range(1, number + 2):
    for col_num in range(1, number + 2):

        # Headers:
        # The first cell(1, 1) is empty.
        if row_num == 1 and col_num == 1:
            sheet.cell(row=row_num, column=col_num).value = ''
        # The first row (upper header in bold) number 1.
        elif row_num == 1:
            sheet.cell(row=row_num, column=col_num).value = col_num - 1  # The second argument of range was + 2.
            sheet.cell(row=row_num, column=col_num).font = bold_font
        # The first column A (leftmost header in bold).
        elif col_num == 1:
            sheet.cell(row=row_num, column=col_num).value = row_num - 1  # The second argument of range was + 2.
            sheet.cell(row=row_num, column=col_num).font = bold_font

        # Finaly calculate the table.
        else:
            sheet.cell(row=row_num, column=col_num).value = (row_num - 1) * (col_num - 1)
            # The second argument of range was 2 more so now it has to be just 1 less.
            # E.g. range(1, 11) goes through 1 to 10, not 11!

home = str(Path.home())
wb.save(f'{home}\\Desktop\\table_{number}.xlsx')
sys.exit(-1)

# Without specifying the exact path, the spreadsheet is saved to the user's folder.

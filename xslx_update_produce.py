#! python3
# xlsx_update_produce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produce_sales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices.
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

for row_num in range(2, sheet.max_row):  # Skip the first row of the header - range(2, 23758). First row = 1 not 0.
    produce_name = sheet.cell(row_num, 1).value  # The second argument is a column. First column = 1 not 0.
    if produce_name in PRICE_UPDATES:
        sheet.cell(row_num, 2).value = PRICE_UPDATES[produce_name]

wb.save('updated_produce_sales.xlsx')
